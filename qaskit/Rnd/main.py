from QuantumHardwareArchitecture import QuantumHardwareArchitecture
import pennylane as qml
from pennylane import numpy as np
from Hamiltonians import *
import random
from QuantumProcess import quantum_processor_launch, objective_function_measurement, \
    gradient_measurement, gradient_decent_one_step, quantum_processor_terminate, set_noise_model
import gc

dev = qml.device("default.qubit", wires=6, shots=1000)
ideal_dev = qml.device("default.qubit", wires=6)


def less_than_num_cx_enum(qubit_num, available_cx: set, cx_num):
    if cx_num == 0:
        return [['I' for _ in range(qubit_num)]]

    cx_state_frame = []
    determined_states = {('I' for _ in range(qubit_num))}
    for pair in available_cx:
        cx_state = ['I' for _ in range(qubit_num)]
        cx_state[pair[0]] = 'C' + str(pair[1])
        cx_state[pair[1]] = 'T' + str(pair[0])
        temp_acx = available_cx.copy()
        waiting_to_discard = set()
        for p in temp_acx:
            if len({p[0], p[1], pair[0], pair[1]}) != 4:
                waiting_to_discard.add(p)
        for p in waiting_to_discard:
            temp_acx.discard(p)
        determined_states.add(tuple(cx_state))
        cx_state_frame.append({'state': cx_state, 'available_cx': temp_acx})

    for current_cx_num in range(2, cx_num + 1):
        extended_frame = []
        for frame in cx_state_frame:
            for pair in frame['available_cx']:
                cx_state = frame['state'].copy()
                waiting_to_discard = set()
                acx = frame['available_cx'].copy()
                cx_state[pair[0]] = 'C' + str(pair[1])
                cx_state[pair[1]] = 'T' + str(pair[0])

                for p in frame['available_cx']:
                    if len({p[0], p[1], pair[0], pair[1]}) != 4:
                        waiting_to_discard.add(p)
                determined_states.add(tuple(cx_state))
                for p in waiting_to_discard:
                    acx.discard(p)
                if len(frame['available_cx']) != 0:
                    extended_frame.append({'state': cx_state, 'available_cx': acx})

        if len(extended_frame) == 0:
            break
        cx_state_frame = extended_frame

    res_list = []
    for tp in determined_states:
        res_list.append(list(tp))
    return res_list


def full_rand_test(qubit_num, Hamiltonian, bond, samples, max_L=3):
    coupling = {(i, np.mod(i + 1, qubit_num)) for i in range(qubit_num)}
    qha = QuantumHardwareArchitecture(
        qml.device("default.mixed", wires=qubit_num, shots=1000), coupling)
    minloss_circ = []
    minloss_par = []
    minloss = 9999
    for cand in range(samples):
        par_num = 0
        circ = []
        for _ in range(max_L):
            single_qubit_gates = [random.choice(['RZ', 'RY', 'I']) for _ in range(qubit_num)]
            cx_gates = random.choice(less_than_num_cx_enum(qubit_num, qha.get_coupling(), int(qubit_num / 2)))
            for q in range(qubit_num):
                if single_qubit_gates[q] != 'I':
                    circ.append((single_qubit_gates[q], q))
                    par_num += 1
            for q in range(qubit_num):
                if cx_gates[q][0] == 'C':
                    target = int(cx_gates[q][1:])
                    circ.append(('Cx', q, target))
        par = np.array([random.uniform(-np.pi, np.pi) for _ in range(par_num)])
        cost_fn = objective_function_measurement(Standard_to_Pennylane_Observables(Hamiltonian, bond), circ)
        cost = cost_fn(par)
        if cost < minloss:
            minloss = cost
            minloss_circ = circ
            minloss_par = par

    record = []
    measurements = samples
    cost_fn = objective_function_measurement(Standard_to_Pennylane_Observables(Hamiltonian, bond), minloss_circ)
    grad_fn = gradient_measurement(Standard_to_Pennylane_Observables(Hamiltonian, bond), minloss_circ)
    par = minloss_par
    par_num = len(par)
    last_learning_rate = 5
    for _ in range(50):
        grad = grad_fn(par)
        measurements += 2 * par_num
        cost = cost_fn(par)
        updated = gradient_decent_one_step(grad, cost, cost_fn, par, learning_rate=last_learning_rate * 2)
        par = updated['parameters']
        measurements += updated['measurements']
        last_learning_rate = updated['learning_rate']
        record.append({'cost': cost_fn(par), 'mea': measurements})
        if last_learning_rate * np.sqrt((grad ** 2).sum()) / par_num <= 0.003:
            break

    result = {'cost': record[-1]['cost'], 'Mea': record[-1]['mea'], 'detail': record}
    return result


def full_rand_multi_trials_test(qubit_num, Hamiltonian, bond, samples, total_trial, pre='', cpus=8):
    from openpyxl import Workbook, load_workbook

    # --------------------------- Prepare the record xlsx file ---------------------------
    set_noise_model({'depolarizing': {'single': 0.001, 'multi': 0.01}})
    quantum_processor_launch(qubit_num, 1000, cpus)
    filename = pre + ('' if pre == '' else '_') + str(bond) + '_Rnd_Multi_Trial_Test.xlsx'
    wb = Workbook()
    ws_final_res = wb.active
    ws_final_res.title = "result"

    col = {'trial': 1, 'cost': 2, 'Mea': 3}

    for key in col:
        ws_final_res.cell(1, col[key], key)

    _ = wb.create_sheet('detail')

    wb.save(filename)

    # --------------------------- Start multi-trail test ---------------------------

    for trial_iter in range(total_trial):
        res = full_rand_test(qubit_num=qubit_num, Hamiltonian=Hamiltonian, bond=str(bond), samples=samples)
        print('trial: ' + str(trial_iter) + ', cost: ' + str(res['cost']) + ', measurements: ' + str(res['Mea']))
        wb = load_workbook(filename)
        ws_final_res = wb['result']
        ws_final_res.cell(trial_iter + 2, col['trial'], trial_iter)
        ws_final_res.cell(trial_iter + 2, col['cost'], float(res['cost']))
        ws_final_res.cell(trial_iter + 2, col['Mea'], res['Mea'])

        ws_detail = wb['detail']
        ws_detail.cell(1, 2 * trial_iter + 1, trial_iter)
        ws_detail.merge_cells(start_row=1, start_column=2 * trial_iter + 1, end_row=1, end_column=2 * trial_iter + 2)
        ws_detail.cell(2, 2 * trial_iter + 1, 'Mea')
        ws_detail.cell(2, 2 * trial_iter + 2, 'cost')

        for rec_iter in range(len(res['detail'])):
            ws_detail.cell(rec_iter + 3, 2 * trial_iter + 1, res['detail'][rec_iter]['mea'])
            ws_detail.cell(rec_iter + 3, 2 * trial_iter + 2, float(res['detail'][rec_iter]['cost']))
        wb.save(filename)
        gc.collect()
    quantum_processor_terminate()


if __name__ == '__main__':
    import sys

    qn = int(sys.argv[1])
    prefix = str(sys.argv[2])
    Hamil = None
    if prefix == 'H2':
        Hamil = H2_Standard
    elif prefix == 'H4':
        Hamil = H4_Standard
    elif prefix == 'HF':
        Hamil = HF_Standard
    else:
        exit(1)
    bnd = str(sys.argv[3])
    samples = int(sys.argv[4])
    trials = int(sys.argv[5])
    proc_num = int(sys.argv[6])

    print('qubit: ' + str(qn))
    print('Hamiltonian: ' + str(prefix))
    print('bond: ' + bnd)
    print('samples: ' + str(samples))
    print('trial: ' + str(trials))
    print('CPUs: ' + str(proc_num))
    full_rand_multi_trials_test(qn, Hamil, bnd, samples, trials, cpus=proc_num)
